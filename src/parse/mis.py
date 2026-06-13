"""Parse the Multi-IsnadSet (MIS) dataset into staging Parquet.

MIS is the platform's dedicated *multiple-isnad* source: a single Sahih Muslim
hadith carries **several parallel isnad chains** (the ``ح`` / *tahwil* split). The
whole value of this source is that those parallel chains are preserved as
distinct transmission paths — collapsing them into one chain would throw away
exactly what MIS exists to model.

Two output files (both tagged ``sect=sunni`` / ``source_corpus=mis``):

* ``hadiths_mis.parquet``        — one ``HADITH_SCHEMA`` row per hadith.
* ``network_edges_mis.parquet``  — ``NETWORK_EDGE_SCHEMA`` directed narrator
  pairs. NETWORK_EDGE is deliberately the right shape here: a hadith's N isnads
  contribute N independent edge-chains that coexist as distinct rows, so the
  multiplicity survives into staging. (The mention schema, keyed only by
  ``source_hadith_id`` + ``position_in_chain`` with no per-chain discriminator,
  would merge the parallel asanid into one chain — the failure mode this parser
  is built to avoid.) The ``from``/``to`` names feed ``make_canonical_id`` via the
  disambiguation/load layer.

Input shape
-----------
Two Mendeley workbooks (``.xlsx``; ``.csv`` exports are also accepted for tests):

* **CoreInfo** — one row per hadith: a hadith number, an (optional) book number,
  the matn text, and an (optional) narrator-sequence string.
* **DetailsInfo (Sanad/Narrators)** — the per-position detail the chains are
  walked from. Two layouts are supported:
    - *sequence layout* (primary): ``(hadith, isnad/sanad, position, narrator)``
      — grouped by ``(hadith, isnad)`` and walked into consecutive-pair edges, so
      each sanad of a hadith becomes its own chain;
    - *edge layout*: explicit ``(from_narrator, to_narrator[, hadith][, isnad])``
      rows — each row is already one edge.

Column names are matched case/spacing/underscore-insensitively (the upstream
spreadsheets are not perfectly normalized), with documented candidate sets.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

import pyarrow as pa

from src.parse.base import generate_source_id, read_csv_robust, safe_int, safe_str, write_parquet
from src.parse.schemas import (
    EDGE_RELATION_TRANSMITTED_TO,
    HADITH_SCHEMA,
    NETWORK_EDGE_SCHEMA,
)
from src.utils.logging import get_logger

logger = get_logger(__name__)

SOURCE_CORPUS = "mis"
SECT = "sunni"
COLLECTION_NAME = "Sahih Muslim"
COLLECTION_SLUG = "sahih_muslim"

# --- column-name candidate sets (matched on the alnum-normalized header) -------
_HADITH_NO_KEYS = ("hadithno", "hadithnumber", "hadithid", "hadith", "serialno", "sno", "no", "id")
_BOOK_KEYS = ("bookno", "booknumber", "book", "bookname", "kitab")
_MATN_KEYS = ("matn", "matntext", "matnarabic", "hadithtext", "text", "arabictext", "matnar")
_ISNAD_TEXT_KEYS = ("narratorsequence", "isnadtext", "sanadtext", "chain", "narrators")

_SANAD_KEYS = (
    "isnadno",
    "isnadid",
    "isnadindex",
    "isnad",
    "sanadno",
    "sanadid",
    "sanad",
    "chainno",
    "chainid",
    "chainindex",
)
_POSITION_KEYS = (
    "position",
    "narratorposition",
    "order",
    "narratororder",
    "sequence",
    "narratorsequence",
    "seq",
    "level",
    "rank",
    "step",
)
_NARRATOR_NAME_KEYS = (
    "narrator",
    "narratorname",
    "narratorarabic",
    "narratorar",
    "name",
    "fullname",
    "rawiname",
)
_NARRATOR_ID_KEYS = ("narratorid", "narratorno", "nid", "rawiid", "rawino", "rawi")
_FROM_KEYS = (
    "fromnarrator",
    "sourcenarrator",
    "from",
    "source",
    "predecessor",
    "teacher",
    "fromname",
)
_TO_KEYS = ("tonarrator", "targetnarrator", "to", "target", "successor", "student", "toname")

# Filename tokens (matched case-insensitively, in priority order) that identify
# each workbook. CoreInfo never contains a DetailsInfo token and vice versa, so
# the two selections never collide.
_DATA_SUFFIXES = (".xlsx", ".xls", ".csv")
_CORE_TOKENS = ("coreinfo", "core")
_DETAIL_TOKENS = ("detailsinfo", "sanad", "narrator", "details")


def _norm_header(value: Any) -> str:
    """Normalize a header to lowercase alphanumerics (drop spaces/underscores)."""
    return "".join(ch for ch in str(value).lower() if ch.isalnum())


def _cell(value: Any) -> str | None:
    """Stringify a spreadsheet/CSV cell, returning None for empties."""
    return safe_str(value)


def _pick(keys: set[str], candidates: tuple[str, ...]) -> str | None:
    """Return the first candidate present in *keys* (already alnum-normalized)."""
    for candidate in candidates:
        if candidate in keys:
            return candidate
    return None


def _read_xlsx_rows(path: Path) -> list[dict[str, str | None]]:
    """Read an .xlsx workbook's first sheet into normalized-header row dicts."""
    from openpyxl import load_workbook

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet = workbook.active or workbook.worksheets[0]
        row_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            return []
        headers = [_norm_header(h) for h in header_row]
        rows: list[dict[str, str | None]] = []
        for raw in row_iter:
            if raw is None or all(cell is None for cell in raw):
                continue
            row = {headers[i]: _cell(raw[i]) for i in range(min(len(headers), len(raw)))}
            rows.append(row)
        return rows
    finally:
        workbook.close()


def _read_csv_rows(path: Path) -> list[dict[str, str | None]]:
    """Read a CSV (robust encoding) into normalized-header row dicts."""
    table, _enc = read_csv_robust(path)
    headers = [_norm_header(c) for c in table.column_names]
    columns = {headers[i]: table.column(i).to_pylist() for i in range(len(headers))}
    return [
        {header: _cell(columns[header][r]) for header in headers} for r in range(table.num_rows)
    ]


def _read_rows(path: Path) -> list[dict[str, str | None]]:
    """Read a CoreInfo/DetailsInfo file (xlsx or csv) into row dicts."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        return _read_xlsx_rows(path)
    return _read_csv_rows(path)


def _find_file(source_dir: Path, tokens: tuple[str, ...]) -> Path | None:
    """Return the first data file under *source_dir* whose name contains a token.

    Case-insensitive (the upstream filenames are CamelCase), token-priority
    ordered, restricted to spreadsheet/CSV suffixes.
    """
    candidates = sorted(
        path
        for path in source_dir.rglob("*")
        if path.is_file() and path.suffix.lower() in _DATA_SUFFIXES
    )
    for token in tokens:
        for path in candidates:
            if token in path.name.lower():
                return path
    return None


def _parse_core(core_path: Path) -> tuple[pa.Table, dict[str, str]]:
    """Parse CoreInfo into a HADITH_SCHEMA table + a ``hadith-no -> source_id`` map.

    The map is the join key the edge builder uses so an edge's ``hadith_id``
    matches its hadith's ``source_id`` exactly (same corpus-namespaced grammar).
    """
    rows = _read_rows(core_path)
    if not rows:
        msg = f"CoreInfo file is empty: {core_path}"
        raise ValueError(msg)

    keys = set(rows[0].keys())
    no_col = _pick(keys, _HADITH_NO_KEYS)
    if no_col is None:
        msg = f"CoreInfo has no hadith-number column. Headers: {sorted(keys)}"
        raise ValueError(msg)
    book_col = _pick(keys, _BOOK_KEYS)
    matn_col = _pick(keys, _MATN_KEYS)
    isnad_text_col = _pick(keys, _ISNAD_TEXT_KEYS)

    hadith_rows: list[dict[str, Any]] = []
    id_map: dict[str, str] = {}
    seen: set[str] = set()

    for row in rows:
        raw_no = row.get(no_col)
        hadith_no = safe_str(raw_no)
        if not hadith_no or hadith_no in seen:
            continue
        seen.add(hadith_no)

        book_num = safe_int(row.get(book_col)) if book_col else None
        source_id = generate_source_id(
            SOURCE_CORPUS, COLLECTION_SLUG, book_num if book_num is not None else 0, hadith_no
        )
        id_map[hadith_no] = source_id

        full_row: dict[str, Any] = {field.name: None for field in HADITH_SCHEMA}
        full_row.update(
            source_id=source_id,
            source_corpus=SOURCE_CORPUS,
            collection_name=COLLECTION_NAME,
            book_number=book_num,
            hadith_number=safe_int(raw_no),
            matn_ar=safe_str(row.get(matn_col)) if matn_col else None,
            isnad_raw_ar=safe_str(row.get(isnad_text_col)) if isnad_text_col else None,
            sect=SECT,
        )
        hadith_rows.append(full_row)

    if not hadith_rows:
        msg = "No valid hadith rows parsed from CoreInfo"
        raise ValueError(msg)

    table = pa.table(
        {field.name: [r[field.name] for r in hadith_rows] for field in HADITH_SCHEMA},
        schema=HADITH_SCHEMA,
    )
    logger.info("mis_core_parsed", hadiths=len(hadith_rows))
    return table, id_map


def _emit_edge(
    edges: list[dict[str, str | None]],
    from_name: str | None,
    to_name: str | None,
    hadith_id: str | None,
    from_id: str | None = None,
    to_id: str | None = None,
) -> None:
    """Append one NETWORK_EDGE row, skipping incomplete / self-loop pairs.

    MIS rows are *isnad transmission* pairs, so each declares the TRANSMITTED_TO
    relation — never STUDIED_UNDER. The graph loader keys on that field, which is
    what keeps these edges off the STUDIED_UNDER loader once MIS is wired into the
    production load path (da#133).
    """
    if not from_name or not to_name or from_name == to_name:
        return
    edges.append(
        {
            "from_narrator_name": from_name,
            "to_narrator_name": to_name,
            "hadith_id": hadith_id,
            "source": SOURCE_CORPUS,
            "from_external_id": from_id,
            "to_external_id": to_id,
            "relation": EDGE_RELATION_TRANSMITTED_TO,
        }
    )


def _parse_edges_sequence(
    rows: list[dict[str, str | None]],
    keys: set[str],
    id_map: dict[str, str],
) -> tuple[list[dict[str, str | None]], dict[str, int]]:
    """Build edges from the per-position *sequence* layout.

    Rows are grouped by ``(hadith, isnad)`` so every sanad of a hadith is walked
    into its own chain — this is where the multiplicity is preserved. Returns
    ``(edges, isnad_count_per_hadith)``.
    """
    hadith_col = _pick(keys, _HADITH_NO_KEYS)
    sanad_col = _pick(keys, _SANAD_KEYS)
    pos_col = _pick(keys, _POSITION_KEYS)
    name_col = _pick(keys, _NARRATOR_NAME_KEYS)
    nid_col = _pick(keys, _NARRATOR_ID_KEYS)

    if hadith_col is None or sanad_col is None or name_col is None:
        msg = (
            "DetailsInfo sequence layout needs hadith + isnad/sanad + narrator columns; "
            f"resolved hadith={hadith_col!r} sanad={sanad_col!r} narrator={name_col!r}. "
            f"Headers: {sorted(keys)}"
        )
        raise ValueError(msg)

    # group key -> list of (position, name, narrator_id), in file order.
    chains: dict[tuple[str, str], list[tuple[int, str | None, str | None]]] = defaultdict(list)
    for idx, row in enumerate(rows):
        hadith_no = safe_str(row.get(hadith_col))
        sanad_no = safe_str(row.get(sanad_col))
        if not hadith_no or not sanad_no:
            continue
        name = safe_str(row.get(name_col))
        nid = safe_str(row.get(nid_col)) if nid_col else None
        pos = safe_int(row.get(pos_col)) if pos_col else None
        # Fall back to file order when no explicit position column.
        order = pos if pos is not None else idx
        chains[(hadith_no, sanad_no)].append((order, name, nid))

    edges: list[dict[str, str | None]] = []
    isnad_counts: dict[str, int] = defaultdict(int)
    for (hadith_no, _sanad_no), members in chains.items():
        isnad_counts[hadith_no] += 1
        ordered = sorted(members, key=lambda m: m[0])
        hadith_id = id_map.get(hadith_no) or generate_source_id(
            SOURCE_CORPUS, COLLECTION_SLUG, 0, hadith_no
        )
        for i in range(len(ordered) - 1):
            _, from_name, from_id = ordered[i]
            _, to_name, to_id = ordered[i + 1]
            _emit_edge(edges, from_name, to_name, hadith_id, from_id, to_id)

    return edges, dict(isnad_counts)


def _parse_edges_explicit(
    rows: list[dict[str, str | None]],
    keys: set[str],
    id_map: dict[str, str],
    from_col: str,
    to_col: str,
) -> tuple[list[dict[str, str | None]], dict[str, int]]:
    """Build edges from the explicit ``(from, to)`` *edge* layout.

    Each row is one edge. ``(hadith, isnad)`` is tracked only to report the
    multiplicity; every row's edge is retained regardless.
    """
    hadith_col = _pick(keys, _HADITH_NO_KEYS)
    sanad_col = _pick(keys, _SANAD_KEYS)

    edges: list[dict[str, str | None]] = []
    seen_isnads: dict[str, set[str]] = defaultdict(set)
    for idx, row in enumerate(rows):
        hadith_no = safe_str(row.get(hadith_col)) if hadith_col else None
        hadith_id = (id_map.get(hadith_no) if hadith_no else None) or (
            generate_source_id(SOURCE_CORPUS, COLLECTION_SLUG, 0, hadith_no) if hadith_no else None
        )
        if hadith_no:
            sanad_no = (safe_str(row.get(sanad_col)) if sanad_col else None) or str(idx)
            seen_isnads[hadith_no].add(sanad_no)
        _emit_edge(edges, safe_str(row.get(from_col)), safe_str(row.get(to_col)), hadith_id)

    isnad_counts = {hadith_no: len(sanads) for hadith_no, sanads in seen_isnads.items()}
    return edges, isnad_counts


def _parse_edges(detail_path: Path, id_map: dict[str, str]) -> pa.Table:
    """Parse DetailsInfo into a NETWORK_EDGE_SCHEMA table, preserving multiplicity."""
    rows = _read_rows(detail_path)
    if not rows:
        msg = f"DetailsInfo file is empty: {detail_path}"
        raise ValueError(msg)

    keys = set(rows[0].keys())
    from_col = _pick(keys, _FROM_KEYS)
    to_col = _pick(keys, _TO_KEYS)

    if from_col and to_col:
        edges, isnad_counts = _parse_edges_explicit(rows, keys, id_map, from_col, to_col)
        layout = "edge"
    else:
        edges, isnad_counts = _parse_edges_sequence(rows, keys, id_map)
        layout = "sequence"

    if not edges:
        msg = "No network edges extracted from MIS DetailsInfo"
        raise ValueError(msg)

    table = pa.table(
        {field.name: [e[field.name] for e in edges] for field in NETWORK_EDGE_SCHEMA},
        schema=NETWORK_EDGE_SCHEMA,
    )

    multi = sum(1 for c in isnad_counts.values() if c > 1)
    logger.info(
        "mis_edges_parsed",
        layout=layout,
        edges=len(edges),
        hadiths=len(isnad_counts),
        multi_isnad_hadiths=multi,
        max_isnad_count=max(isnad_counts.values()) if isnad_counts else 0,
    )
    return table


def run(raw_dir: Path, staging_dir: Path) -> tuple[Path, Path]:
    """Parse the MIS workbooks into hadith + network-edge Parquet files."""
    source_dir = raw_dir / "mis"
    if not source_dir.exists():
        msg = f"Source directory not found: {source_dir}"
        raise FileNotFoundError(msg)

    core_path = _find_file(source_dir, _CORE_TOKENS)
    if core_path is None:
        msg = f"No MIS CoreInfo file found under {source_dir}"
        raise FileNotFoundError(msg)
    detail_path = _find_file(source_dir, _DETAIL_TOKENS)
    if detail_path is None:
        msg = f"No MIS DetailsInfo (sanad/narrators) file found under {source_dir}"
        raise FileNotFoundError(msg)

    logger.info("mis_files_found", core=str(core_path), detail=str(detail_path))

    hadith_table, id_map = _parse_core(core_path)
    hadiths_path = write_parquet(
        hadith_table, staging_dir / "hadiths_mis.parquet", schema=HADITH_SCHEMA
    )

    edge_table = _parse_edges(detail_path, id_map)
    edges_path = write_parquet(
        edge_table, staging_dir / "network_edges_mis.parquet", schema=NETWORK_EDGE_SCHEMA
    )

    return hadiths_path, edges_path
